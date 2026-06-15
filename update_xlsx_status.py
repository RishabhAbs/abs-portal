"""Mark completed tasks in Boss_Tasks_List.xlsx with Status='Done' + green fill + strikethrough."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

DONE_TASKS = {
    "Customer Inactive Option only for Admin",
    "user not allow to change there target",
    "Expiry Report Export option only Visiable with Admin",
    "User Button Join with Profile",
    "Top Dashboard Header Fix",
    "Customer Search Mobile Responsive",
    "Mobile no is Inactive but search Work 9706050760",
    "user filter in target Report",
    "Tally API Update Button in Search",
    "Our Serial Update Button Inactive",
    "Create a seprate Option for Master IN this Add Item",
}

xlsx_path = r"c:\Users\hp\Downloads\abscloud\abscloud\Boss_Tasks_List_updated.xlsx"
wb = load_workbook(xlsx_path)
ws = wb["By Priority"]

done_fill = PatternFill("solid", fgColor="C6EFCE")
done_font = Font(bold=True, color="006100", size=10, strike=True)
status_font = Font(bold=True, color="006100", size=10)

thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

updated = 0
for row in range(2, ws.max_row + 1):
    task_cell = ws.cell(row=row, column=2)
    if not task_cell.value:
        continue
    task = str(task_cell.value).strip()
    if task in DONE_TASKS:
        # Strikethrough + green fill across the row
        for col in range(1, 11):
            c = ws.cell(row=row, column=col)
            c.fill = done_fill
            if col == 10:  # Status column
                c.value = "Done"
                c.font = status_font
            else:
                c.font = done_font
        updated += 1
        print(f"  Marked done: {task}")

print(f"\nUpdated {updated} tasks.")
try:
    wb.save(xlsx_path)
    print(f"Saved: {xlsx_path}")
except PermissionError:
    out = xlsx_path.replace(".xlsx", "_updated.xlsx")
    wb.save(out)
    print(f"File was locked. Saved to: {out}")
