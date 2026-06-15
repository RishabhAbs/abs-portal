"""
Build Boss_Tasks_List.xlsx organized by priority tier.

Tiers:
  QUICK WIN  - <= 4 hrs, single-file or single-flag changes, low risk
  MEDIUM     - 0.5 to 2 days, new reports/feature additions, multi-file
  BIG / WOW  - multi-day; new modules, integrations, infra
  INVESTIGATE- bug reports needing root-cause analysis before scoping

Each task is tagged with: tier, effort (XS/S/M/L/XL), impact (Low/Med/High),
suggested order, breakdown (sub-tasks or what-to-check), risks, owner, status.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# (task, original_type, tier, effort, impact, breakdown, risks_or_notes)
TASKS = [
    # =================== QUICK WINS (knock out first for momentum) ===================
    ("Customer Inactive Option only for Admin", "REQUIREMENT", "Quick Win", "XS", "Med",
     "Wrap toggle in role check (admin only) on UI + API guard.", "None - simple role gate."),
    ("user filter in target Report", "CORRECTION", "Quick Win", "XS", "Med",
     "Add user dropdown filter; pass user_id to query.", "Make sure default is 'All' for admins."),
    ("user not allow to change there target", "CORRECTION", "Quick Win", "XS", "Med",
     "Make target field read-only for non-admin role.", "Confirm which roles can edit (admin/manager)."),
    ("Expiry Report Export option only Visiable with Admin", "CORRECTION", "Quick Win", "XS", "Low",
     "Conditional render Export button based on role.", "Same role-gate pattern."),
    ("User Button Join with Profile", "CORRECTION", "Quick Win", "XS", "Low",
     "Wrap user button in router-link to /profile/:id.", "Confirm profile route exists."),
    ("Lead report with last update detail", "CORRECTION", "Quick Win", "XS", "Med",
     "Add 'Last Updated' column from existing updated_at field.", "May need join if stored separately."),
    ("Top Dashboard Header Fix", "CORRECTION", "Quick Win", "XS", "Low",
     "CSS fix on dashboard header (alignment / sticky / overflow).", "Verify across screen sizes."),
    ("Customer Search Mobile Responsive", "REQUIREMENT", "Quick Win", "S", "High",
     "Add responsive CSS breakpoints for search results table.", "Test on actual mobile, not just resized browser."),
    ("Tally API Update Button in Search", "CORRECTION", "Quick Win", "S", "Med",
     "Add 'Update from Tally' button in search row; call existing tally hit endpoint.", "Depends on Tally API task #1; may block."),
    ("Our Serial Update Button Inactive", "CORRECTION", "Quick Win", "XS", "Low",
     "Fix disabled state logic on Serial Update button.", "Find why it's stuck disabled - likely a flag check."),
    ("Create a seprate Option for Master IN this Add Item", "CORRECTION", "Quick Win", "S", "Med",
     "Add 'Master' radio/toggle in Add Item form; route to master table.", "Confirm UX with boss before building."),

    # =================== MEDIUM (proper feature work) ===================
    ("Outstanding Report", "REQUIREMENT", "Medium", "M", "High",
     "1) DB query (party-wise outstanding)  2) API endpoint  3) UI table + filters  4) Export PDF/Excel", "Define grouping: by party / by office / by aging bucket?"),
    ("Stock Summary Report Required", "REQUIREMENT", "Medium", "M", "High",
     "1) Aggregate stock by item/category  2) API  3) UI with filters (date, item)  4) Export", "Confirm whether opening+closing or only current."),
    ("New Party Ledger Report", "REQUIREMENT", "Medium", "M", "High",
     "1) Ledger query (debit/credit per party)  2) API  3) UI with date range  4) PDF export", "Compare against existing party report - avoid duplicate."),
    ("DashBoard Devide Sales and Outstanding, Customer", "REQUIREMENT", "Medium", "M", "High",
     "Split dashboard into 3 cards/sections: Sales / Outstanding / Customers; add navigation drill-down.", "Get layout mock from boss before coding."),
    ("Category wise Amount and Quantity Define for Target", "CORRECTION", "Medium", "M", "Med",
     "Extend target schema to per-category; UI to set qty + amount per category; rollup logic.", "Migration if schema changes - check existing data."),
    ("When Click expiry report then auto expiry update", "CORRECTION", "Medium", "S", "Med",
     "On expiry-report load, trigger expiry status recompute (or scheduled job + manual trigger).", "Heavy queries - run async if dataset large."),
    ("Pending Report 2 option Required till Today Pending", "CORRECTION", "Medium", "S", "Med",
     "Add 2nd pending option = pending up to today (vs current 'all pending').", "Clarify naming with boss to avoid confusion."),
    ("External Task Report With Last Update Report", "CORRECTION", "Medium", "M", "Med",
     "Add last_update column + filter to external task report; backend join with task_updates.", "Index on updated_at for performance."),
    ("Task Report Change Based on External Task", "CORRECTION", "Medium", "M", "Med",
     "Modify task report to reflect external task linkage; UI + query change.", "Define behavior when external task is deleted."),
    ("Search Lastcall, Last Visit and Service Call popup", "CORRECTION", "Medium", "M", "Med",
     "Popup component showing last call / last visit / service call history per customer in search.", "Decide modal vs hover-card UX."),
    ("Correction and requirement edit by the developer", "REQUIREMENT", "Medium", "M", "Med",
     "Allow developer role to edit submitted correction/requirement entries (with audit trail).", "Audit log mandatory - boss will want history."),

    # =================== BIG / WOW (multi-day, multi-component) ===================
    ("Vouchers Tally API", "REQUIREMENT", "Big / Wow", "XL", "High",
     "1) Auth with Tally  2) Voucher schema mapping  3) Push voucher endpoint  4) Pull voucher endpoint  5) Error/retry handling  6) Logging  7) Admin UI to monitor sync", "Tally XML format quirks; offline queue if Tally is down."),
    ("Tally api hit and Maintain Record for Updated Data", "REQUIREMENT", "Big / Wow", "XL", "High",
     "1) Track updated_at per entity  2) Delta query API  3) Scheduled job (cron) to call Tally  4) Sync log table  5) Conflict resolution  6) Retry on failure  7) Dashboard for sync status", "Idempotency critical - duplicate writes will corrupt Tally."),
    ("Cloud Billing System", "REQUIREMENT", "Big / Wow", "XL", "High",
     "1) Pricing/plan model  2) Subscription table  3) Invoice generator  4) Payment gateway integration  5) Customer billing UI  6) Admin reports  7) Webhooks for payment events", "Major scope - get spec/wireframes before estimating dates."),
    ("DataBase Auto Backup", "REQUIREMENT", "Big / Wow", "L", "High",
     "1) mysqldump cron (daily + weekly)  2) S3/local rotated storage  3) Restore tested  4) Email/Slack alert on failure  5) Admin UI to download backup", "Test RESTORE, not just backup. Untested backup = no backup."),
    ("add Complain SEction", "REQUIREMENT", "Big / Wow", "L", "High",
     "1) Complaint table + status enum  2) CRUD APIs  3) Customer-facing form  4) Admin queue UI  5) Assign-to-user  6) Status workflow  7) Email/SMS notifications", "Define SLA/status flow with boss before schema."),

    # =================== INVESTIGATE (root-cause first, scope unclear) ===================
    ("Attendenace Report Error", "CORRECTION", "Investigate", "?", "Med",
     "REPRO: which user/date triggers the error? Check logs, query, null fields. Likely fix: S-M.", "Don't patch - find the actual bug; could indicate bad data."),
    ("Billing Daybook Amount Error and Required Period", "CORRECTION", "Investigate", "?", "High",
     "REPRO with sample dates. Check sum/aggregate query. Add period filter once amount issue is isolated.", "Amount errors = customer-trust bug; reproduce with real numbers."),
    ("Mobile no is Inactive but search Work 9706050760", "CORRECTION", "Investigate", "?", "Med",
     "Why is an inactive mobile returning in search? Check WHERE clause - missing is_active=1?", "May be intentional for admin - confirm before fixing."),
    ("Check out Lead Auto Buissness Closed", "CORRECTION", "Investigate", "?", "Med",
     "Investigate auto-close logic for Lead/Business; what trigger fires it? Is rule correct?", "Could be a cron or status-change trigger - audit both."),
    ("Search Update issue", "CORRECTION", "Investigate", "?", "Med",
     "Vague - get repro steps from boss/Vansh. What 'update' fails? In which screen?", "Don't guess - clarify scope before estimating."),
]

# Sanity check
assert len(TASKS) == 32, f"Expected 32 tasks, got {len(TASKS)}"

TIER_CONFIG = {
    "Quick Win":   {"fill": "70AD47", "order": 1, "subtitle": "Knock these out first - momentum + clear backlog fast"},
    "Medium":      {"fill": "FFC000", "order": 2, "subtitle": "Standard feature work - half day to two days each"},
    "Big / Wow":   {"fill": "C00000", "order": 3, "subtitle": "Major modules - need design + spec before coding"},
    "Investigate": {"fill": "5B9BD5", "order": 4, "subtitle": "Root-cause first; scope/effort unknown until reproduced"},
}

EFFORT_HINT = {
    "XS": "<= 2h",
    "S":  "2-4h",
    "M":  "0.5-2d",
    "L":  "3-5d",
    "XL": "1-2 wk",
    "?":  "TBD",
}

# ----------------- Build workbook -----------------
wb = Workbook()

# Common styles
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1F4E78")
section_font = Font(bold=True, color="FFFFFF", size=12)

req_fill = PatternFill("solid", fgColor="DDEBF7")
corr_fill = PatternFill("solid", fgColor="FFF2CC")

impact_fills = {
    "High": PatternFill("solid", fgColor="F4B084"),
    "Med":  PatternFill("solid", fgColor="FFE699"),
    "Low":  PatternFill("solid", fgColor="C6E0B4"),
}

# ============ Sheet 1: By Priority (master view) ============
ws = wb.active
ws.title = "By Priority"

headers = ["#", "Task", "Original Type", "Effort", "Effort Hint", "Impact",
           "Breakdown / Sub-steps", "Risks / Notes", "Owner", "Status"]
ws.append(headers)
for col_idx in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border
ws.row_dimensions[1].height = 30

row = 2
sno = 1

# Group tasks by tier in defined order
for tier_name, cfg in sorted(TIER_CONFIG.items(), key=lambda x: x[1]["order"]):
    tier_tasks = [t for t in TASKS if t[2] == tier_name]
    if not tier_tasks:
        continue
    # Section banner
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
    c = ws.cell(row=row, column=1, value=f"  {tier_name.upper()}  ({len(tier_tasks)})  -  {cfg['subtitle']}")
    c.font = section_font
    c.fill = PatternFill("solid", fgColor=cfg["fill"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24
    row += 1

    for task, otype, _tier, effort, impact, breakdown, risks in tier_tasks:
        values = [
            sno, task, otype, effort, EFFORT_HINT.get(effort, ""), impact,
            breakdown, risks, "", "Pending",
        ]
        for col_idx, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.border = border
            if col_idx in (2, 7, 8):
                c.alignment = left
            else:
                c.alignment = center
            if col_idx == 3:
                c.fill = req_fill if otype == "REQUIREMENT" else corr_fill
                c.font = Font(bold=True, size=9)
            if col_idx == 6 and impact in impact_fills:
                c.fill = impact_fills[impact]
                c.font = Font(bold=True, size=9)
        sno += 1
        row += 1

# Column widths
widths = [5, 48, 14, 8, 10, 8, 60, 45, 14, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# ============ Sheet 2: Suggested Sprint Order ============
ws2 = wb.create_sheet("Suggested Order")
ws2.append(["Day / Slot", "Tier", "Task", "Effort", "Why now"])
for col_idx in range(1, 6):
    c = ws2.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

ORDER = [
    # Day 1-2: clear quick wins
    ("Day 1 AM", "Quick Win", "Top Dashboard Header Fix", "XS", "5-min CSS - start with a guaranteed win"),
    ("Day 1 AM", "Quick Win", "Our Serial Update Button Inactive", "XS", "Likely a flag check - quick fix"),
    ("Day 1 AM", "Quick Win", "User Button Join with Profile", "XS", "Just a router-link - trivial"),
    ("Day 1 PM", "Quick Win", "Customer Inactive Option only for Admin", "XS", "Single role-gate pattern - reuse it"),
    ("Day 1 PM", "Quick Win", "Expiry Report Export option only Visiable with Admin", "XS", "Same pattern - cheap to ship together"),
    ("Day 1 PM", "Quick Win", "user not allow to change there target", "XS", "Same role-gate pattern"),
    ("Day 2 AM", "Quick Win", "user filter in target Report", "XS", "Add dropdown + query param"),
    ("Day 2 AM", "Quick Win", "Lead report with last update detail", "XS", "Add column from existing field"),
    ("Day 2 PM", "Quick Win", "Customer Search Mobile Responsive", "S", "Customer-facing - high impact"),
    ("Day 2 PM", "Quick Win", "Create a seprate Option for Master IN this Add Item", "S", "Confirm UX first, then build"),
    ("Day 3 AM", "Investigate", "Search Update issue", "?", "Get clarity from boss/Vansh - blocks others"),
    ("Day 3 AM", "Investigate", "Mobile no is Inactive but search Work 9706050760", "?", "Likely 1-line WHERE fix once reproduced"),
    ("Day 3 PM", "Investigate", "Attendenace Report Error", "?", "Reproduce + log dive"),
    # Day 4+: medium feature work
    ("Day 4-5", "Medium", "Outstanding Report", "M", "High impact - boss will see results"),
    ("Day 4-5", "Medium", "Stock Summary Report Required", "M", "Pair with Outstanding - similar plumbing"),
    ("Day 6-7", "Medium", "New Party Ledger Report", "M", "Third report - reuse export code"),
    ("Day 8", "Medium", "DashBoard Devide Sales and Outstanding, Customer", "M", "Visible to boss every login"),
    ("Day 9", "Medium", "Pending Report 2 option Required till Today Pending", "S", "Quick once context loaded"),
    ("Day 9", "Medium", "When Click expiry report then auto expiry update", "S", "Logic tweak"),
    ("Day 10", "Medium", "Search Lastcall, Last Visit and Service Call popup", "M", "Popup UX"),
    ("Day 11", "Medium", "External Task Report With Last Update Report", "M", "Pair with #22"),
    ("Day 11", "Medium", "Task Report Change Based on External Task", "M", "Same area as #21"),
    ("Day 12", "Medium", "Category wise Amount and Quantity Define for Target", "M", "Schema change - do mid-sprint"),
    ("Day 13", "Medium", "Correction and requirement edit by the developer", "M", "Audit-log heavy"),
    ("Day 14", "Investigate", "Billing Daybook Amount Error and Required Period", "?", "Customer-money bug - debug carefully"),
    ("Day 14", "Investigate", "Check out Lead Auto Buissness Closed", "?", "Audit auto-close trigger"),
    ("Day 14", "Quick Win", "Tally API Update Button in Search", "S", "Best done after Tally Big tasks have a working endpoint"),
    # Big tasks - parallel track / dedicated weeks
    ("Wk 3", "Big / Wow", "DataBase Auto Backup", "L", "Lowest-spec Big - ship first to de-risk"),
    ("Wk 3", "Big / Wow", "add Complain SEction", "L", "Self-contained module"),
    ("Wk 4-5", "Big / Wow", "Vouchers Tally API", "XL", "Foundation for Tally sync"),
    ("Wk 4-5", "Big / Wow", "Tally api hit and Maintain Record for Updated Data", "XL", "Builds on Vouchers Tally API"),
    ("Wk 6-7", "Big / Wow", "Cloud Billing System", "XL", "Largest scope - needs spec from boss before starting"),
]

for r_idx, (slot, tier, task, effort, why) in enumerate(ORDER, 2):
    ws2.cell(row=r_idx, column=1, value=slot).alignment = center
    ws2.cell(row=r_idx, column=2, value=tier).alignment = center
    tier_cell = ws2.cell(row=r_idx, column=2)
    tier_cell.fill = PatternFill("solid", fgColor=TIER_CONFIG[tier]["fill"])
    tier_cell.font = Font(bold=True, color="FFFFFF", size=9)
    ws2.cell(row=r_idx, column=3, value=task).alignment = left
    ws2.cell(row=r_idx, column=4, value=effort).alignment = center
    ws2.cell(row=r_idx, column=5, value=why).alignment = left
    for col_idx in range(1, 6):
        ws2.cell(row=r_idx, column=col_idx).border = border

for i, w in enumerate([12, 14, 50, 8, 55], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# ============ Sheet 3: Big/Wow Breakdown (one page per major task) ============
ws3 = wb.create_sheet("Big-Wow Breakdown")
ws3.append(["Big Task", "Sub-step #", "Sub-step", "Done?"])
for col_idx in range(1, 5):
    c = ws3.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

BIG_BREAKDOWN = [
    ("Vouchers Tally API", [
        "Set up Tally connector lib + auth/handshake",
        "Map our Voucher schema -> Tally XML format",
        "Push voucher endpoint (POST /api/tally/voucher)",
        "Pull voucher endpoint (GET /api/tally/vouchers?since=...)",
        "Error handling + retry queue (failed pushes)",
        "Sync log table + admin UI to view status",
        "Test with real Tally instance (not mock)",
    ]),
    ("Tally api hit and Maintain Record for Updated Data", [
        "Add updated_at column to all syncable entities (if missing)",
        "Build delta-query API (return rows updated since timestamp)",
        "Cron job: every N minutes, push deltas to Tally",
        "Sync_log table: entity, last_synced_at, status, error",
        "Idempotency: dedupe key so retries don't double-write",
        "Conflict resolution rules (latest-wins? manual?)",
        "Admin dashboard - sync health per entity",
    ]),
    ("Cloud Billing System", [
        "Spec/wireframes from boss - DO THIS FIRST",
        "Plans + pricing tables (DB schema)",
        "Subscription model linking customer -> plan",
        "Invoice generator (monthly cron + on-demand)",
        "Payment gateway integration (Razorpay / Stripe / etc.)",
        "Customer billing UI (view invoices, pay)",
        "Admin reports - MRR, overdue, churn",
        "Webhooks for payment success/failure",
    ]),
    ("DataBase Auto Backup", [
        "mysqldump script with daily + weekly rotation",
        "Storage location (S3 bucket OR external disk) + access keys",
        "Cron entry on server (verify it runs)",
        "Restore drill - prove the backup actually works",
        "Email/Slack alert on backup failure",
        "Admin UI to list + download backups",
        "Document the recovery procedure",
    ]),
    ("add Complain SEction", [
        "Spec with boss - SLA, status flow, who-sees-what",
        "DB schema: complaints + complaint_updates + status enum",
        "CRUD API endpoints",
        "Customer-facing submission form",
        "Admin queue UI - filter by status / assignee",
        "Assign-to-user + reassignment workflow",
        "Status transitions (Open -> In Progress -> Resolved -> Closed)",
        "Email / SMS / WhatsApp notifications on status change",
    ]),
]

row = 2
for big, steps in BIG_BREAKDOWN:
    start_row = row
    for i, step in enumerate(steps, 1):
        ws3.cell(row=row, column=1, value=big if i == 1 else "")
        ws3.cell(row=row, column=2, value=i).alignment = center
        ws3.cell(row=row, column=3, value=step).alignment = left
        ws3.cell(row=row, column=4, value="").alignment = center
        for col_idx in range(1, 5):
            ws3.cell(row=row, column=col_idx).border = border
        row += 1
    # Merge big task name across its rows
    if len(steps) > 1:
        ws3.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
    head_cell = ws3.cell(row=start_row, column=1)
    head_cell.font = Font(bold=True, size=11, color="FFFFFF")
    head_cell.fill = PatternFill("solid", fgColor="C00000")
    head_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # spacer row
    row += 1

for i, w in enumerate([28, 10, 70, 8], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A2"

# ============ Sheet 4: Legend / Key ============
ws4 = wb.create_sheet("Legend")
legend = [
    ["Field", "Values", "Meaning"],
    ["Tier", "Quick Win", "<=4h work, low risk - do first"],
    ["", "Medium", "Half-day to 2-day standard feature work"],
    ["", "Big / Wow", "Multi-day major module - needs spec first"],
    ["", "Investigate", "Bug needs reproduction before scoping"],
    ["Effort", "XS", "<= 2 hours"],
    ["", "S",  "2 to 4 hours"],
    ["", "M",  "0.5 to 2 days"],
    ["", "L",  "3 to 5 days"],
    ["", "XL", "1 to 2 weeks"],
    ["", "?",  "Cannot estimate yet (Investigate tier)"],
    ["Impact", "High", "Customer-facing OR boss watches it OR money/data correctness"],
    ["", "Med", "Internal team productivity / quality of life"],
    ["", "Low", "Polish or rare-edge fix"],
    ["Original Type", "REQUIREMENT", "New feature requested"],
    ["", "CORRECTION", "Bug / fix to existing feature"],
]
for r in legend:
    ws4.append(r)
for col_idx in range(1, 4):
    c = ws4.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border
for r_idx in range(2, len(legend) + 1):
    for c_idx in range(1, 4):
        cell = ws4.cell(row=r_idx, column=c_idx)
        cell.border = border
        cell.alignment = left if c_idx == 3 else center
for i, w in enumerate([15, 14, 60], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

out = r"c:\Users\hp\Downloads\abscloud\abscloud\Boss_Tasks_List.xlsx"
wb.save(out)

# Stats
from collections import Counter
tier_counts = Counter(t[2] for t in TASKS)
print("Saved:", out)
print("Tasks per tier:")
for tier in ["Quick Win", "Medium", "Big / Wow", "Investigate"]:
    print(f"  {tier:<14}: {tier_counts[tier]}")
print("Total:", sum(tier_counts.values()))
